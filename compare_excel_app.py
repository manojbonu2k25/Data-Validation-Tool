import streamlit as st
import pandas as pd
import io
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

st.set_page_config(page_title="Excel Data Validation", page_icon="📊", layout="centered")
st.image("vector.png", use_container_width=100)
st.write("## Data Validation Tool")
# col1, col2 = st.columns([1,3])
# with col1:
#     st.image("vector.png", use_container_width=200)
# with col2:
#     st.write("### Data Validation Tool")
st.write("Upload two Excel files, select comparison type, and view results.")

# === Upload files ===
file_a = st.file_uploader("📂 Upload First Excel File", type=["xlsx", "xls"])
file_b = st.file_uploader("📂 Upload Second Excel File", type=["xlsx", "xls"])

# === Comparison Type Dropdown ===
comparison_type = st.selectbox(
    "Select Type of Comparison:",
    ["Sheet Difference", "Stacked Comparison", "Calculation Difference"]
)

if file_a and file_b:
    excel_a = pd.ExcelFile(io.BytesIO(file_a.getvalue()))
    excel_b = pd.ExcelFile(io.BytesIO(file_b.getvalue()))

    sheet_a = st.selectbox(f"Select sheet from **{file_a.name}**:", excel_a.sheet_names, key="sheet_a_select")
    sheet_b = st.selectbox(f"Select sheet from **{file_b.name}**:", excel_b.sheet_names, key="sheet_b_select")

    df_a = pd.read_excel(io.BytesIO(file_a.getvalue()), sheet_name=sheet_a)
    df_b = pd.read_excel(io.BytesIO(file_b.getvalue()), sheet_name=sheet_b)

    # =============================================
    # 1️⃣ SHEET DIFFERENCE MODE
    # =============================================
    if comparison_type == "Sheet Difference":
        st.subheader("🔹 Sheet Difference Mode")
        cols_a = st.multiselect("Columns from Sheet A:", df_a.columns.tolist(), default=df_a.columns.tolist(), key="cols_a")
        cols_b = st.multiselect("Columns from Sheet B:", df_b.columns.tolist(), default=df_b.columns.tolist(), key="cols_b")

        if len(cols_a) != len(cols_b):
            st.warning("⚠️ Please select the same number of columns in both sheets.")
        else:
            if st.button("🔍 Run Sheet Difference"):
                df_a_sel = df_a[cols_a].copy()
                df_b_sel = df_b[cols_b].copy()
                df_b_sel.columns = df_a_sel.columns

                df_a_sel = df_a_sel.drop_duplicates()
                df_b_sel = df_b_sel.drop_duplicates()
                df_a_sel = df_a_sel.astype(str).apply(lambda x: x.str.strip())
                df_b_sel = df_b_sel.astype(str).apply(lambda x: x.str.strip())

                merged = df_a_sel.merge(df_b_sel, how='outer', indicator=True)
                not_in_b = merged[merged['_merge'] == 'left_only'].drop('_merge', axis=1)
                not_in_a = merged[merged['_merge'] == 'right_only'].drop('_merge', axis=1)

                # 🧾 Summary Counts
                total_a = len(df_a_sel)
                total_b = len(df_b_sel)
                only_in_b = len(not_in_b)
                only_in_a = len(not_in_a)
                matched = total_a + total_b - (only_in_a + only_in_b)

                st.markdown("### 📈 Summary")
                st.write(f"**Total Rows in Sheet A:** {total_a}")
                st.write(f"**Total Rows in Sheet B:** {total_b}")
                st.write(f"✅ **Matched Rows:** {matched}")
                st.write(f"🔵 **Rows in A not in B:** {only_in_b}")
                st.write(f"🟡 **Rows in B not in A:** {only_in_a}")

                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    not_in_b.to_excel(writer, sheet_name="In_A_Not_in_B", index=False)
                    not_in_a.to_excel(writer, sheet_name="In_B_Not_in_A", index=False)
                output.seek(0)

                st.success("✅ Sheet Difference completed!")
                st.download_button(
                    label="📥 Download Result Excel",
                    data=output,
                    file_name="SheetDifference_Result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("📊 Preview of Differences")
                st.dataframe(not_in_b if not not_in_b.empty else pd.DataFrame(["No differences found."], columns=["Message"]))
                st.dataframe(not_in_a if not not_in_a.empty else pd.DataFrame(["No differences found."], columns=["Message"]))

    # =============================================
    # 2️⃣ STACKED COMPARISON MODE
    # =============================================
    elif comparison_type == "Stacked Comparison":
        st.subheader("🔹 Stacked Comparison Mode")

        name_a = st.text_input("Enter a name for File A (e.g. CDK):", "File A")
        name_b = st.text_input("Enter a name for File B (e.g. AXEL):", "File B")
        control_col = st.text_input("Enter Control Column Name (case-insensitive):", "Control number")

        if st.button("🔍 Run Stacked Comparison"):
            def find_col_case_insensitive(df, target):
                for c in df.columns:
                    if str(c).strip().lower() == target.strip().lower():
                        return c
                raise KeyError(f"Column '{target}' not found (case-insensitive).")

            control_a = find_col_case_insensitive(df_a, control_col)
            control_b = find_col_case_insensitive(df_b, control_col)
            df_a = df_a.rename(columns={control_a: control_col}).copy()
            df_b = df_b.rename(columns={control_b: control_col}).copy()

            df_a[control_col] = df_a[control_col].astype(str).str.strip()
            df_b[control_col] = df_b[control_col].astype(str).str.strip()

            df_a["Source"] = name_a
            df_b["Source"] = name_b

            all_cols = list(dict.fromkeys(
                ["Source", control_col]
                + [c for c in df_a.columns if c not in ("Source", control_col)]
                + [c for c in df_b.columns if c not in ("Source", control_col)]
            ))
            df_a = df_a.reindex(columns=all_cols)
            df_b = df_b.reindex(columns=all_cols)

            a_set = set(df_a[control_col].dropna().astype(str))
            b_set = set(df_b[control_col].dropna().astype(str))
            paired_set = a_set & b_set
            only_in_a_set = a_set - b_set
            only_in_b_set = b_set - a_set

            stacked = pd.concat([df_a, df_b], ignore_index=True)

            def pair_status(ctrl, src):
                if ctrl in paired_set:
                    return "Paired"
                if src == name_a and ctrl in only_in_a_set:
                    return f"{name_a}-only"
                if src == name_b and ctrl in only_in_b_set:
                    return f"{name_b}-only"
                return ""

            stacked["PairStatus"] = [pair_status(str(r[control_col]), r["Source"]) for _, r in stacked.iterrows()]

            # Sorting to group A–B pairs together
            def sort_key(row):
                ctrl = str(row[control_col])
                src = row["Source"]
                block = 0 if ctrl in paired_set else (1 if ctrl in only_in_a_set else 2)
                src_order = 0 if src == name_a else 1
                return (block, ctrl, src_order)

            stacked = stacked.loc[sorted(stacked.index, key=lambda i: sort_key(stacked.loc[i]))].reset_index(drop=True)

            # 🧾 Summary Counts
            total_a = len(df_a)
            total_b = len(df_b)
            total_paired = len(paired_set)
            only_a = len(only_in_a_set)
            only_b = len(only_in_b_set)

            st.markdown("### 📈 Summary")
            st.write(f"**Total Rows in {name_a}:** {total_a}")
            st.write(f"**Total Rows in {name_b}:** {total_b}")
            st.write(f"✅ **Paired Records:** {total_paired}")
            st.write(f"🔵 **{name_a}-only Records:** {only_a}")
            st.write(f"🟡 **{name_b}-only Records:** {only_b}")

            # === Excel output with highlighting ===
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                stacked.to_excel(writer, sheet_name="Combined", index=False)
                df_a[df_a[control_col].isin(only_in_a_set)].to_excel(writer, sheet_name=f"{name_a}-only", index=False)
                df_b[df_b[control_col].isin(only_in_b_set)].to_excel(writer, sheet_name=f"{name_b}-only", index=False)
            output.seek(0)

            wb = load_workbook(output)
            ws = wb["Combined"]
            headers = {cell.value: idx + 1 for idx, cell in enumerate(ws[1])}
            source_idx = headers["Source"]
            ctrl_idx = headers[control_col]

            YELLOW = PatternFill(start_color="FFF4B3", end_color="FFF4B3", fill_type="solid")  # File B
            BLUE = PatternFill(start_color="CFE2F3", end_color="CFE2F3", fill_type="solid")    # File A paired

            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"

            for r in range(2, ws.max_row + 1):
                src = (ws.cell(r, source_idx).value or "").strip().upper()
                ctrl = str(ws.cell(r, ctrl_idx).value).strip().upper()
                if src == name_b.upper():
                    fill = YELLOW
                elif src == name_a.upper() and ctrl in (c.upper() for c in paired_set):
                    fill = BLUE
                else:
                    fill = None
                if fill:
                    for c in range(1, ws.max_column + 1):
                        ws.cell(r, c).fill = fill

            out_mem = io.BytesIO()
            wb.save(out_mem)
            out_mem.seek(0)

            st.success("✅ Stacked Comparison completed!")
            st.download_button(
                label="📥 Download Stacked Result Excel",
                data=out_mem,
                file_name="StackedComparison_Result.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )

            st.subheader("📊 Preview of Combined Data (Top 50 Rows)")
            st.dataframe(stacked.head(50))

    # =============================================
    # 3️⃣ CALCULATION DIFFERENCE MODE (new option)
    # =============================================
    elif comparison_type == "Calculation Difference":
        st.subheader("🔹 Calculation Difference Mode")
        st.write("Compare numeric columns between File A and File B using a common key.")

        # Get sheet selections
        name_a = st.text_input("Enter a name for File A (e.g. CDK):", "File A")
        name_b = st.text_input("Enter a name for File B (e.g. AXEL):", "File B")

        control_col = st.selectbox("Select matching key column (common between both files):", df_a.columns)
        num_col_a = st.selectbox(f"Select numeric column from {name_a}:", df_a.columns)
        num_col_b = st.selectbox(f"Select numeric column from {name_b}:", df_b.columns)

        if st.button("🔍 Run Calculation Difference"):
            try:
                # Prepare datasets
                df_a_tmp = df_a[[control_col, num_col_a]].copy()
                df_b_tmp = df_b[[control_col, num_col_b]].copy()

                df_a_tmp[control_col] = df_a_tmp[control_col].astype(str).str.strip()
                df_b_tmp[control_col] = df_b_tmp[control_col].astype(str).str.strip()

                df_a_tmp[num_col_a] = pd.to_numeric(df_a_tmp[num_col_a], errors="coerce")
                df_b_tmp[num_col_b] = pd.to_numeric(df_b_tmp[num_col_b], errors="coerce")

                # Merge and compute difference
                merged = pd.merge(df_a_tmp, df_b_tmp, on=control_col, how="inner")
                merged["Difference"] = merged[num_col_a] - merged[num_col_b]

                # Summary
                st.markdown("### 📈 Summary")
                st.write(f"**Total Matched Rows:** {len(merged)}")
                st.write(f"**Average Difference:** {merged['Difference'].mean():.2f}")
                st.write(f"**Positive Differences (A>B):** {(merged['Difference']>0).sum()}")
                st.write(f"**Negative Differences (A<B):** {(merged['Difference']<0).sum()}")

                # Export Excel
                output = io.BytesIO()
                with pd.ExcelWriter(output, engine="openpyxl") as writer:
                    merged.to_excel(writer, sheet_name="Differences", index=False)
                output.seek(0)

                st.success("✅ Calculation Difference completed!")
                st.download_button(
                    label="📥 Download Result Excel",
                    data=output,
                    file_name="CalculationDifference_Result.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                st.subheader("📊 Preview of Differences (Top 50 Rows)")
                st.dataframe(merged.head(50))

            except Exception as e:
                st.error(f"Error while computing differences: {e}")